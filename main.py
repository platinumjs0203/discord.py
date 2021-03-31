import discord
import os
import asyncio
import random
import datetime
import time
from bs4 import BeautifulSoup
import requests
from urllib.request import urlopen, Request
import urllib
import urllib.request
from openpyxl import load_workbook

from discord.ext.commands import Bot

client = discord.Client()
game = discord.Game("인간들에게 복수를 준비")


headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36'}


def create_soup(url):
    res = requests.get(url, headers=headers)
    res.raise_for_status()
    html = res.text
    soup = BeautifulSoup(html, 'html.parser')
    return soup

# ========== 토큰암호화 ===========


token_path = os.path.dirname(os.path.abspath(__file__)) + '/token.txt'
t = open(token_path, 'r', encoding='utf-8')
token = t.read().split()[0]

#  ============ 가동 =============


@client.event
async def on_ready():
    print("시스템 가동 준비완료")
    print(client.user.name)
    print("----------------------")
    await client.change_presence(activity=game)
    await client.get_channel(792704016697917462).send('내가 돌아왔다.')


@client.event
async def on_message(message):
    if message.content.startswith("$도움"):
        dtime = datetime.datetime.now()
        year = dtime.year
        month = dtime.month
        day = dtime.day
        embed = discord.Embed(
            title='명령어 정보',
            description=f'=========== {year}년 {month}월 {day}일 명령어 모음 ===========',
            colour=discord.Colour.gold()
        )
        await message.channel.send(embed=embed)

    # ============= 실시간 뉴스 검색 ===============

    if message.content.startswith("$뉴스"):
        url = 'https://news.naver.com'
        soup = create_soup(url)
        headline = soup.find(
            'ul', 'hdline_article_list').find_all('li', limit=3)
        await message.channel.send('현재시각 주요뉴스를 알려드립니다.')
        for news in headline:
            link = url + news.find('a')['href']
            await message.channel.send(link)

    # ============= 날 씨 ===============

    if message.content.startswith("$날씨"):
        try:
            search = message.content.split("/")
            loca = search[1]
            search_loca = urllib.parse.quote(loca + "날씨")
            url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=1&ie=utf8&query=" + search_loca
            soup = create_soup(url)
            cast = soup.find(
                'p', attrs={'class': 'cast_txt'}).get_text().replace('아요', '')
            curr_temp = soup.find(
                'p', attrs={'class': 'info_temperature'}).get_text().replace('도씨', '')
            min_temp = soup.find('span', attrs={'class': 'min'}).get_text()
            max_temp = soup.find('span', attrs={'class': 'max'}).get_text()

            morning_rain_rate = soup.find(
                'span', attrs={'class': 'point_time morning'}).get_text().strip()
            afternoon_rain_rate = soup.find(
                'span', attrs={'class': 'point_time afternoon'}).get_text().strip()

            dust = soup.find('dl', attrs={'class': 'indicator'})
            pm10 = dust.find_all('dd')[0].get_text()
            pm25 = dust.find_all('dd')[1].get_text()

            embed = discord.Embed(
                title=f'현재 {loca} 날씨 입니다.',
                description='=====================',
                colour=discord.Colour.gold()
            )
            embed.add_field(
                name='현재온도', value=f'현재 {curr_temp} (최저 {min_temp} / 최고 {max_temp})\n{cast}습니다.', inline=False)
            embed.add_field(
                name='강수확률', value=f'오전 {morning_rain_rate}\n오후 {afternoon_rain_rate}', inline=False)
            embed.add_field(name='미세먼지', value=f'{pm10}\n{pm25}')
            await message.channel.send(embed=embed)
        except IndexError:
            embed = discord.Embed(
                title="에러!! 지역을 입력해주세요!!",
                description="혹시 $날씨 만 입력하셨거나 /를 입력안하신건 아닌가요?\nex)$날씨/부산 이라고 입력해보세요!",
                colour=discord.Colour.red()
            )
            await message.channel.send(embed=embed)
        except AttributeError:
            embed = discord.Embed(
                title="에러!! 정확한 지명을 입력해주세요!!",
                description="정확한 지명을 입력해주세요!!\nex)$날씨/부산명지 ❌ $날씨/부산명지동 ⭕",
                colour=discord.Colour.red()
            )
            await message.channel.send(embed=embed)

    # ============= 포켓몬 뉴스 ===============

    if message.content.startswith("$포켓몬고/뉴스"):
        url = "https://gugomah.tistory.com/"
        soup = create_soup(url)
        poke_table = soup.find("div", "cover-thumbnail-2")
        poke_news = poke_table.find_all("li", limit=3)
        for i in poke_news:
            link = url + i.find('a')['href']
            await message.channel.send(link)

    # ============= 랜덤 고양이 ===============

    if message.content.startswith("$짤"):
        wb = load_workbook('cat_image.xlsx')
        ws = wb['Sheet1']
        random_url = range(0, 1474)
        random_img = random.choice(random_url)
        zzal = ws[f'A{random_img}'].value
        embed = discord.Embed(
            title='귀여운 고양이를 드리겠읍니다 ~ 😻', colour=discord.Colour.red())
        await message.channel.send(embed=embed)
        await message.channel.send(zzal)

    # ============== 비트코인 ==============
    if message.content.startswith('$비트코인'):

        while True:
            url = 'https://www.bithumb.com/'
            soup = create_soup(url)
            bitcoin = soup.find(
                'strong', attrs={'id': 'assetRealBTC_KRW'}).get_text()
            bc_y = soup.find('strong', attrs={
                'id': 'assetRealPriceBTC_KRW'}).get_text()
            embed = discord.Embed(
                title=f"현재 비트코인 가격은 {bitcoin} 입니다.",
                description=f"전일대비 {bc_y} 입니다.",
                colour=discord.Colour.gold()
            )
            await message.channel.send(embed=embed)
            await asyncio.sleep(300)


client.run(token)
